import openpyxl
import sys
import re
import app.views 
from app.models import networks, devices
from app import db

def select_types(description): 
	''' проверяет описание на наличие типовых слов и возврашает тип если что-то найдено'''
	if description is None: 
		return None
	description = description.lower() 
	print(description)
	if 'ettop' in description  or 'еттоп' in description: 
		return 'nettop'
	elif 'phone'in description  or 'елефон' in description: 
		return 'telephone'
	elif 'мфу' in description or 'mfp' in description: 
		return 'mfu'
	elif 'онкий' in description or 'ТК' in description: 
		return 'tk'
	elif 'оутер' in description or 'mikrotik' in description:
		return 'router'
	else:
		return None

def range_ip_parce(ws, network, start_col = 2): 
	''' парсим список ip адресов и извлекаем данные об оборудовании '''
	for row in range(9,263): 
		ip = ws.cell(column = start_col, row = row).value
		if ip is not None: 
			#
			description = ws.cell(column = start_col+1, row = row).value
			user = ws.cell(column = start_col+2, row = row).value
			number = ws.cell(column = start_col+3, row = row).value
			comment = str()
			for col in range(start_col+4,start_col+7):	
				val = ws.cell(column = col, row = row).value
				if val is not None:
					comment = comment + ' ' + val
			if number is not None:	
				#оставляем в номере только цифры				
				#number = re.sub('[Ии]нв(\.)|(\.\s)|(\s)','',str(number))
				try:
					number = re.search('\d+',str(number)).group(0)
				except AttributeError:
					pass

			device = devices.query.filter_by(ip = ip).first()
			device.type = select_types(description) 
			device.description = description
			device.comment = comment
			device.number = number 
			device.owner = user
			print (device.ip, device.type)	 
	try:
		db.session.commit() 	
	except: 
		return False
	return True


def excel_parcing(file_path):
	''' парсим данные о сетях (имя, cidr, первый адрес) и вызываем уже парсер ip адресов'''
	wb = openpyxl.load_workbook(filename = file_path)
	for sheet in wb:
		ws = sheet
		#забираем данные из ячеек имя подсети - имя листа
		try:
			cidr = ws['B7'].value[-2::1]
			net = ws['B9'].value
			name = sheet.title
		except TypeError as e:
			print (e)
			continue
		print(name, net, cidr)
		netw = networks(name = name,
            description = name,
            cidr = cidr,
            net = net)
		if len(networks.query.filter(networks.name.like('%' + str(name) + '%')).all()) !=0 : 
			print('Network with name ', name, 'is exist!')
			continue
		
		if app.views.create_network(netw):
			print('Network create ok!')
		else:
			print('Network create error!!')
			return False
		if range_ip_parce(ws, netw): 
			print('IP range ok!')
		else: 
 			print('IP range error!')
 			return False
 		#если подсеть 23 то парсим еще ячейки с 10 по 19
		if cidr == '23':
			range_ip_parce(ws,netw,start_col = 10)
	return True	



	 





